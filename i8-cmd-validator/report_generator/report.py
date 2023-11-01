
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class Report:
    def __init__(self, path) -> None:
        self.path = path
        self.column_names = [
            "Index",
            "Board",
            "Command",
            "Command Target",
            "Frame Start",
            "Length",
            "Command",
            "Checksum 1",
            "Data",
            "Checksum 2",
            "Frame End",
            "Full Command",
            "Response",
            "Response Status"
        ]
        self.workbook, self.worksheet = self.setup()
    
    def setup(self):
        if Path(self.path).is_file():
            workbook = load_workbook(self.path)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            self.add_column_names(worksheet)
        return workbook, worksheet

    def add_column_names(self, worksheet):
        font = Font(bold=True)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center')
        for col_num, column_title in enumerate(self.column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = font
            cell.border = border
            cell.alignment = alignment
            cell.value = column_title

    def append(self, row):
        if not isinstance(row, list):
            raise ValueError("Row must be a list")
        # Añadir el índice al principio de la fila
        row_with_index = [self.worksheet.max_row] + row
        self.worksheet.append(row_with_index)

    def add_bold_separator(self):
        last_row = self.worksheet.max_row
        bold_border = Border(bottom=Side(style='thick'))
        # Colocar el separador en la última fila escrita, sin añadir una nueva fila
        for col_num in range(1, len(self.column_names) + 1):  # Empezar desde la segunda columna
            cell = self.worksheet.cell(row=last_row, column=col_num)
            cell.border = bold_border

    def export(self):
        self.worksheet.freeze_panes = 'A2'  # Fijar la fila de los nombres de las columnas
        self.workbook.save(self.path)